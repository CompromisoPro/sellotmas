# -*- coding: utf-8 -*-
"""
Genera templates COMPLETOS para consultores basados en los 58 criterios
y 12 tipos de empresa del SelloPRO.
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from collections import defaultdict

# ============================================
# Colores
# ============================================
GREEN_FILL = PatternFill(start_color="35D38B", end_color="35D38B", fill_type="solid")
DARK_FILL = PatternFill(start_color="343434", end_color="343434", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
NORMAL_FONT = Font(size=9)
BOLD_FONT = Font(bold=True, size=9)
TITLE_FONT = Font(bold=True, size=14, color="343434")
THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

# ============================================
# Cargar datos del CSV
# ============================================
def cargar_datos_cuestionarios():
    archivo = r"c:\Users\acarreno\Downloads\Base Sello pro, Herramientas.xlsx - Base Formularios.csv"
    
    criterios = defaultdict(lambda: {
        'ambito': '',
        'estrellas': {},
        'tipos_empresa': set()
    })
    
    tipos_empresa = []
    
    with open(archivo, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        tipos_empresa = header[3:]
        
        for row in reader:
            if len(row) < 4:
                continue
            
            ambito = row[0].strip()
            criterio = ' '.join(row[1].strip().replace('\n', ' ').split())
            
            if not ambito or not criterio:
                continue
            
            try:
                estrella = int(row[2].strip())
            except:
                continue
            
            key = (ambito, criterio)
            criterios[key]['ambito'] = ambito
            
            if estrella not in criterios[key]['estrellas']:
                criterios[key]['estrellas'][estrella] = {}
            
            for i, tipo in enumerate(tipos_empresa):
                if i + 3 < len(row):
                    desc = row[i + 3].strip()
                    if desc and desc.lower() != 'x':
                        criterios[key]['estrellas'][estrella][tipo] = desc
                        criterios[key]['tipos_empresa'].add(tipo)
    
    return dict(criterios), tipos_empresa


def crear_template_matriz_completa():
    """
    Crea el template de Matriz de Vinculacion con TODOS los 58 criterios.
    """
    print("Cargando datos de cuestionarios...")
    criterios, tipos_empresa = cargar_datos_cuestionarios()
    print(f"  - {len(criterios)} criterios encontrados")
    print(f"  - {len(tipos_empresa)} tipos de empresa")
    
    wb = openpyxl.Workbook()
    
    # ========== HOJA 1: INSTRUCCIONES ==========
    ws_instr = wb.active
    ws_instr.title = "INSTRUCCIONES"
    
    instrucciones = [
        ("INSTRUCCIONES PARA LLENAR LA MATRIZ DE VINCULACION", TITLE_FONT, None),
        ("", None, None),
        ("CONTEXTO:", BOLD_FONT, None),
        ("Esta matriz establece la relacion entre cada criterio del SelloPRO y los objetivos", NORMAL_FONT, None),
        ("de la Taxonomia de Actividades Medioambientalmente Sostenibles de Chile (T-MAS).", NORMAL_FONT, None),
        ("", None, None),
        ("IMPORTANTE - HAY 58 CRITERIOS EN TOTAL:", BOLD_FONT, None),
        ("Los criterios varian segun el tipo de empresa (Grande vs Pequena/mediana) y sector.", NORMAL_FONT, None),
        ("Algunos criterios NO aplican a ciertos tipos de empresa (marcados con 'x' en el original).", NORMAL_FONT, None),
        ("", None, None),
        ("QUE NECESITAMOS DE USTEDES:", BOLD_FONT, None),
        ("1. VALIDAR el tipo de vinculo (Directo/Indirecto/Sin Vinculo) para CADA criterio", NORMAL_FONT, None),
        ("2. INDICAR desde que estrella (1-5) aplica el vinculo", NORMAL_FONT, None),
        ("3. MARCAR que objetivos T-MAS impacta cada criterio (SMS, CS_OM1, CS_OM2, NHDS_OM1-6)", NORMAL_FONT, None),
        ("4. ESCRIBIR la justificacion de por que se relaciona (lenguaje simple)", NORMAL_FONT, None),
        ("5. ESCRIBIR como la empresa demuestra el cumplimiento", NORMAL_FONT, None),
        ("6. Para INDIRECTOS: que accion concreta necesita para ser DIRECTO", NORMAL_FONT, None),
        ("", None, None),
        ("HOJAS EN ESTE ARCHIVO:", BOLD_FONT, None),
        ("- Hoja 'Matriz Vinculacion': Los 58 criterios para llenar vinculo T-MAS", NORMAL_FONT, None),
        ("- Hoja 'Estrellas por Tipo': Descripcion de estrellas segun tipo de empresa", NORMAL_FONT, None),
        ("- Hoja 'Resumen Tipos': Que criterios aplican a que tipo de empresa", NORMAL_FONT, None),
        ("", None, None),
        ("DEFINICIONES:", BOLD_FONT, None),
        ("- Directo: El criterio cumple DIRECTAMENTE con el requisito T-MAS", NORMAL_FONT, None),
        ("- Indirecto: Contribuye parcialmente, hay acciones adicionales para cumplir", NORMAL_FONT, None),
        ("- Sin Vinculo: No tiene relacion con la T-MAS", NORMAL_FONT, None),
        ("", None, None),
        ("OBJETIVOS T-MAS:", BOLD_FONT, None),
        ("- SMS = Salvaguardas Minimas Sociales (DDHH, etica, canal de denuncias)", NORMAL_FONT, None),
        ("- CS_OM1 = Contribucion Sustancial a Mitigacion del Cambio Climatico", NORMAL_FONT, None),
        ("- CS_OM2 = Contribucion Sustancial a Adaptacion al Cambio Climatico", NORMAL_FONT, None),
        ("- NHDS_OM1 = No Hacer Dano Significativo a Mitigacion", NORMAL_FONT, None),
        ("- NHDS_OM2 = No Hacer Dano Significativo a Adaptacion", NORMAL_FONT, None),
        ("- NHDS_OM3 = No Hacer Dano Significativo al Uso de Agua", NORMAL_FONT, None),
        ("- NHDS_OM4 = No Hacer Dano Significativo a Economia Circular (Residuos)", NORMAL_FONT, None),
        ("- NHDS_OM5 = No Hacer Dano Significativo a Prevencion de Contaminacion", NORMAL_FONT, None),
        ("- NHDS_OM6 = No Hacer Dano Significativo a Proteccion de Biodiversidad", NORMAL_FONT, None),
        ("", None, None),
        ("PLAZO: Por favor devolver este archivo llenado dentro de 1 semana.", BOLD_FONT, None),
    ]
    
    ws_instr.column_dimensions['A'].width = 90
    for i, (texto, font, fill) in enumerate(instrucciones):
        cell = ws_instr[f'A{i+1}']
        cell.value = texto
        if font:
            cell.font = font
    
    # ========== HOJA 2: MATRIZ VINCULACION ==========
    ws = wb.create_sheet("Matriz Vinculacion")
    
    # Titulo
    ws.merge_cells('A1:Q1')
    ws['A1'] = 'MATRIZ DE VINCULACION SELLOPRO - T-MAS (58 CRITERIOS)'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:Q2')
    ws['A2'] = 'Celdas amarillas = requieren validacion. Celdas naranjas = INDIRECTOS que necesitan "Que falta para Directo"'
    ws['A2'].font = Font(size=9, italic=True, color="92400E")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = [
        ('A', '#', 4),
        ('B', 'Ambito', 18),
        ('C', 'Criterio SelloPRO', 40),
        ('D', 'Tipo\nVinculo', 12),
        ('E', 'Estrella\nMinima', 10),
        ('F', 'SMS', 6),
        ('G', 'CS\nOM1', 6),
        ('H', 'CS\nOM2', 6),
        ('I', 'NHDS\nOM1', 6),
        ('J', 'NHDS\nOM2', 6),
        ('K', 'NHDS\nOM3', 6),
        ('L', 'NHDS\nOM4', 6),
        ('M', 'NHDS\nOM5', 6),
        ('N', 'NHDS\nOM6', 6),
        ('O', 'Justificacion\n(Por que se relaciona)', 45),
        ('P', 'Como lo demuestra\nla empresa', 40),
        ('Q', 'SOLO INDIRECTOS:\nQue falta para Directo', 45),
    ]
    
    row = 4
    for col_letter, header, width in headers:
        cell = ws[f'{col_letter}{row}']
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = DARK_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[row].height = 40
    
    # Datos - ordenar por ambito
    criterios_ordenados = sorted(criterios.items(), key=lambda x: (x[0][0], x[0][1]))
    
    for i, ((ambito, criterio), data) in enumerate(criterios_ordenados):
        r = row + 1 + i
        
        # Numero
        ws[f'A{r}'] = i + 1
        ws[f'A{r}'].font = BOLD_FONT
        ws[f'A{r}'].border = THIN_BORDER
        ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Ambito
        ws[f'B{r}'] = ambito
        ws[f'B{r}'].font = NORMAL_FONT
        ws[f'B{r}'].border = THIN_BORDER
        ws[f'B{r}'].alignment = Alignment(vertical='center', wrap_text=True)
        
        # Criterio
        ws[f'C{r}'] = criterio
        ws[f'C{r}'].font = BOLD_FONT
        ws[f'C{r}'].border = THIN_BORDER
        ws[f'C{r}'].alignment = Alignment(vertical='center', wrap_text=True)
        
        # Tipo vinculo (a llenar)
        ws[f'D{r}'] = ''
        ws[f'D{r}'].fill = YELLOW_FILL
        ws[f'D{r}'].border = THIN_BORDER
        
        # Estrella minima (a llenar)
        ws[f'E{r}'] = ''
        ws[f'E{r}'].fill = YELLOW_FILL
        ws[f'E{r}'].border = THIN_BORDER
        ws[f'E{r}'].alignment = Alignment(horizontal='center')
        
        # Objetivos T-MAS (a llenar)
        for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{r}'] = ''
            ws[f'{col}{r}'].fill = YELLOW_FILL
            ws[f'{col}{r}'].border = THIN_BORDER
            ws[f'{col}{r}'].alignment = Alignment(horizontal='center')
        
        # Justificacion (a llenar)
        ws[f'O{r}'] = ''
        ws[f'O{r}'].fill = YELLOW_FILL
        ws[f'O{r}'].border = THIN_BORDER
        ws[f'O{r}'].alignment = Alignment(vertical='center', wrap_text=True)
        
        # Como demuestra (a llenar)
        ws[f'P{r}'] = ''
        ws[f'P{r}'].fill = YELLOW_FILL
        ws[f'P{r}'].border = THIN_BORDER
        ws[f'P{r}'].alignment = Alignment(vertical='center', wrap_text=True)
        
        # Que falta para directo (naranja - solo para indirectos)
        ws[f'Q{r}'] = ''
        ws[f'Q{r}'].fill = ORANGE_FILL
        ws[f'Q{r}'].border = THIN_BORDER
        ws[f'Q{r}'].alignment = Alignment(vertical='center', wrap_text=True)
        
        ws.row_dimensions[r].height = 45
    
    # Validaciones de datos
    last_row = row + len(criterios_ordenados)
    
    dv_tipo = DataValidation(type="list", formula1='"Directo,Indirecto,Sin Vinculo"', allow_blank=True)
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f'D5:D{last_row}')
    
    dv_estrellas = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
    ws.add_data_validation(dv_estrellas)
    dv_estrellas.add(f'E5:E{last_row}')
    
    dv_sino = DataValidation(type="list", formula1='"Si,No"', allow_blank=True)
    ws.add_data_validation(dv_sino)
    for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        dv_sino.add(f'{col}5:{col}{last_row}')
    
    # ========== HOJA 3: ESTRELLAS POR TIPO ==========
    ws_est = wb.create_sheet("Estrellas por Tipo")
    
    ws_est.merge_cells('A1:N1')
    ws_est['A1'] = 'DESCRIPCION DE ESTRELLAS POR TIPO DE EMPRESA'
    ws_est['A1'].font = TITLE_FONT
    ws_est['A1'].alignment = Alignment(horizontal='center')
    ws_est.row_dimensions[1].height = 30
    
    ws_est.merge_cells('A2:N2')
    ws_est['A2'] = 'Referencia: que significa cada nivel de estrellas segun el tipo de empresa'
    ws_est['A2'].font = Font(size=9, italic=True)
    ws_est['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    est_headers = [
        ('A', '#', 4),
        ('B', 'Ambito', 15),
        ('C', 'Criterio', 35),
        ('D', 'Estrella', 8),
        ('E', 'Constructora-Inmob (G)', 35),
        ('F', 'Constructora-Inmob (P/M)', 35),
        ('G', 'Constructoras (G)', 35),
        ('H', 'Constructoras (P/M)', 35),
        ('I', 'Inmobiliarias (G)', 35),
        ('J', 'Inmobiliarias (P/M)', 35),
        ('K', 'Especialidades (G)', 35),
        ('L', 'Especialidades (P/M)', 35),
        ('M', 'Industriales (G)', 35),
        ('N', 'Industriales (P/M)', 35),
        ('O', 'Proveedores (G)', 35),
        ('P', 'Proveedores (P/M)', 35),
    ]
    
    row_est = 4
    for col_letter, header, width in est_headers:
        cell = ws_est[f'{col_letter}{row_est}']
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = DARK_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws_est.column_dimensions[col_letter].width = width
    ws_est.row_dimensions[row_est].height = 35
    
    # Datos de estrellas
    row_idx = row_est + 1
    criterio_num = 0
    for (ambito, criterio), data in criterios_ordenados:
        criterio_num += 1
        for estrella in sorted(data['estrellas'].keys()):
            ws_est[f'A{row_idx}'] = criterio_num
            ws_est[f'A{row_idx}'].border = THIN_BORDER
            ws_est[f'A{row_idx}'].alignment = Alignment(horizontal='center')
            
            ws_est[f'B{row_idx}'] = ambito
            ws_est[f'B{row_idx}'].border = THIN_BORDER
            ws_est[f'B{row_idx}'].font = NORMAL_FONT
            
            ws_est[f'C{row_idx}'] = criterio
            ws_est[f'C{row_idx}'].border = THIN_BORDER
            ws_est[f'C{row_idx}'].font = NORMAL_FONT
            ws_est[f'C{row_idx}'].alignment = Alignment(wrap_text=True)
            
            ws_est[f'D{row_idx}'] = estrella
            ws_est[f'D{row_idx}'].border = THIN_BORDER
            ws_est[f'D{row_idx}'].alignment = Alignment(horizontal='center')
            
            # Descripciones por tipo
            descripciones = data['estrellas'][estrella]
            col_idx = 4  # Columna E
            for tipo in tipos_empresa:
                col_letter = get_column_letter(col_idx + 1)
                desc = descripciones.get(tipo, 'x')
                ws_est[f'{col_letter}{row_idx}'] = desc
                ws_est[f'{col_letter}{row_idx}'].border = THIN_BORDER
                ws_est[f'{col_letter}{row_idx}'].font = Font(size=8)
                ws_est[f'{col_letter}{row_idx}'].alignment = Alignment(wrap_text=True, vertical='center')
                if desc == 'x':
                    ws_est[f'{col_letter}{row_idx}'].fill = LIGHT_GRAY_FILL
                col_idx += 1
            
            ws_est.row_dimensions[row_idx].height = 50
            row_idx += 1
    
    # ========== HOJA 4: RESUMEN TIPOS ==========
    ws_res = wb.create_sheet("Resumen Tipos")
    
    ws_res.merge_cells('A1:N1')
    ws_res['A1'] = 'QUE CRITERIOS APLICAN A CADA TIPO DE EMPRESA'
    ws_res['A1'].font = TITLE_FONT
    ws_res['A1'].alignment = Alignment(horizontal='center')
    
    ws_res.merge_cells('A2:N2')
    ws_res['A2'] = 'X = NO aplica a ese tipo de empresa'
    ws_res['A2'].font = Font(size=9, italic=True)
    ws_res['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    res_headers = ['#', 'Ambito', 'Criterio'] + [t.replace('Constructora - Inmobiliaria', 'Const-Inmob').replace('Pequeña / mediana', 'P/M').replace('pequeña / mediana', 'P/M').replace('(Grande)', '(G)') for t in tipos_empresa]
    
    row_res = 4
    for col_idx, header in enumerate(res_headers):
        col_letter = get_column_letter(col_idx + 1)
        cell = ws_res[f'{col_letter}{row_res}']
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = DARK_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        if col_idx == 0:
            ws_res.column_dimensions[col_letter].width = 4
        elif col_idx == 1:
            ws_res.column_dimensions[col_letter].width = 15
        elif col_idx == 2:
            ws_res.column_dimensions[col_letter].width = 35
        else:
            ws_res.column_dimensions[col_letter].width = 10
    ws_res.row_dimensions[row_res].height = 50
    
    # Datos
    row_idx = row_res + 1
    for i, ((ambito, criterio), data) in enumerate(criterios_ordenados):
        ws_res[f'A{row_idx}'] = i + 1
        ws_res[f'A{row_idx}'].border = THIN_BORDER
        ws_res[f'A{row_idx}'].alignment = Alignment(horizontal='center')
        
        ws_res[f'B{row_idx}'] = ambito
        ws_res[f'B{row_idx}'].border = THIN_BORDER
        ws_res[f'B{row_idx}'].font = NORMAL_FONT
        
        ws_res[f'C{row_idx}'] = criterio
        ws_res[f'C{row_idx}'].border = THIN_BORDER
        ws_res[f'C{row_idx}'].font = NORMAL_FONT
        ws_res[f'C{row_idx}'].alignment = Alignment(wrap_text=True)
        
        for col_idx, tipo in enumerate(tipos_empresa):
            col_letter = get_column_letter(col_idx + 4)
            aplica = tipo in data['tipos_empresa']
            ws_res[f'{col_letter}{row_idx}'] = 'OK' if aplica else 'X'
            ws_res[f'{col_letter}{row_idx}'].border = THIN_BORDER
            ws_res[f'{col_letter}{row_idx}'].alignment = Alignment(horizontal='center')
            if aplica:
                ws_res[f'{col_letter}{row_idx}'].fill = LIGHT_GREEN_FILL
            else:
                ws_res[f'{col_letter}{row_idx}'].fill = LIGHT_GRAY_FILL
                ws_res[f'{col_letter}{row_idx}'].font = Font(color='94A3B8', size=9)
        
        ws_res.row_dimensions[row_idx].height = 30
        row_idx += 1
    
    # Guardar
    wb.save('TEMPLATE_COMPLETO_vinculacion_tmas.xlsx')
    print("Archivo guardado: TEMPLATE_COMPLETO_vinculacion_tmas.xlsx")


def crear_template_objetivos_tmas():
    """
    Crea template para info oficial de objetivos T-MAS
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Objetivos T-MAS"
    
    ws.merge_cells('A1:K1')
    ws['A1'] = 'INFORMACION OFICIAL DE OBJETIVOS T-MAS'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:K2')
    ws['A2'] = 'Descripciones, requisitos y documentos. VALIDAR con fuentes oficiales de la T-MAS.'
    ws['A2'].font = Font(size=9, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = [
        ('A', 'Codigo', 10),
        ('B', 'Nombre', 25),
        ('C', 'Descripcion\n(lenguaje simple)', 50),
        ('D', 'Requisito 1', 35),
        ('E', 'Requisito 2', 35),
        ('F', 'Requisito 3', 35),
        ('G', 'Requisito 4', 35),
        ('H', 'Documento 1', 25),
        ('I', 'Documento 2', 25),
        ('J', 'Documento 3', 25),
        ('K', 'Referencia Oficial', 30),
    ]
    
    row = 4
    for col_letter, header, width in headers:
        cell = ws[f'{col_letter}{row}']
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = DARK_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = width
    
    objetivos = [
        {"id": "SMS", "nombre": "Salvaguardas Minimas Sociales"},
        {"id": "CS_OM1", "nombre": "Contribucion Sustancial - Mitigacion"},
        {"id": "CS_OM2", "nombre": "Contribucion Sustancial - Adaptacion"},
        {"id": "NHDS_OM1", "nombre": "No Dano Significativo - Mitigacion"},
        {"id": "NHDS_OM2", "nombre": "No Dano Significativo - Adaptacion"},
        {"id": "NHDS_OM3", "nombre": "No Dano Significativo - Agua"},
        {"id": "NHDS_OM4", "nombre": "No Dano Significativo - Residuos"},
        {"id": "NHDS_OM5", "nombre": "No Dano Significativo - Contaminacion"},
        {"id": "NHDS_OM6", "nombre": "No Dano Significativo - Biodiversidad"},
    ]
    
    for i, obj in enumerate(objetivos):
        r = row + 1 + i
        ws[f'A{r}'] = obj['id']
        ws[f'A{r}'].font = BOLD_FONT
        ws[f'A{r}'].border = THIN_BORDER
        
        ws[f'B{r}'] = obj['nombre']
        ws[f'B{r}'].font = BOLD_FONT
        ws[f'B{r}'].border = THIN_BORDER
        ws[f'B{r}'].alignment = Alignment(wrap_text=True)
        
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws[f'{col}{r}'] = ''
            ws[f'{col}{r}'].fill = YELLOW_FILL
            ws[f'{col}{r}'].border = THIN_BORDER
            ws[f'{col}{r}'].alignment = Alignment(wrap_text=True, vertical='center')
        
        ws.row_dimensions[r].height = 60
    
    wb.save('TEMPLATE_objetivos_tmas.xlsx')
    print("Archivo guardado: TEMPLATE_objetivos_tmas.xlsx")


if __name__ == '__main__':
    print("=" * 70)
    print("  GENERANDO TEMPLATES COMPLETOS PARA CONSULTORES")
    print("=" * 70)
    print()
    
    crear_template_matriz_completa()
    crear_template_objetivos_tmas()
    
    print()
    print("=" * 70)
    print("  ARCHIVOS GENERADOS:")
    print("=" * 70)
    print("  1. TEMPLATE_COMPLETO_vinculacion_tmas.xlsx")
    print("     - Hoja 'Instrucciones': Como llenar el template")
    print("     - Hoja 'Matriz Vinculacion': 58 criterios para vincular con T-MAS")
    print("     - Hoja 'Estrellas por Tipo': Descripcion de estrellas x tipo empresa")
    print("     - Hoja 'Resumen Tipos': Que criterios aplican a que tipo")
    print()
    print("  2. TEMPLATE_objetivos_tmas.xlsx")
    print("     - Info oficial de cada objetivo T-MAS")
    print()
    print("=" * 70)


"""
Genera los templates Excel que deben llenar los consultores.
Incluye datos pre-llenados con lo que ya tenemos para que solo validen/corrijan.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================
# Colores
# ============================================
GREEN_FILL = PatternFill(start_color="35D38B", end_color="35D38B", fill_type="solid")
DARK_FILL = PatternFill(start_color="343434", end_color="343434", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
NORMAL_FONT = Font(size=10)
BOLD_FONT = Font(bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="343434")
THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

# ============================================
# Datos actuales (pre-llenados para validación)
# ============================================
CRITERIOS_ACTUALES = [
    {"id": 1, "dim": "Gobernanza", "criterio": "Estrategia de sostenibilidad", "tipo": "Directo", "estrella": 2,
     "sms": "Sí", "cs1": "Sí", "cs2": "Sí", "n1": "Sí", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Tu estrategia de sostenibilidad considera los mismos temas que pide la Taxonomía.",
     "demo": "Mostrando en tu estrategia cómo abordas las Salvaguardas Sociales y la medición de GEI.",
     "falta": ""},
    {"id": 2, "dim": "Gobernanza", "criterio": "Gobierno corporativo", "tipo": "Indirecto", "estrella": 1,
     "sms": "Sí", "cs1": "Sí", "cs2": "Sí", "n1": "Sí", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Los conocimientos del directorio pueden considerar requisitos de la Taxonomía, pero no los cubre directamente.",
     "demo": "Respaldando que los asesores tienen conocimientos sobre la Ley Marco de Cambio Climático.",
     "falta": "El directorio debe capacitarse en los CTS de la T-MAS y documentar cómo los incorporan."},
    {"id": 3, "dim": "Gobernanza", "criterio": "Gestión de la ética e integridad", "tipo": "Indirecto", "estrella": 3,
     "sms": "Sí", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Cumple con una parte de las SMS desde las 3 estrellas.",
     "demo": "Canal de denuncias y código de ética.",
     "falta": "Procedimientos para gestionar riesgos de corrupción y código de ética en toda la cadena de valor."},
    {"id": 4, "dim": "Trabajadores", "criterio": "Conozco a mis trabajadores", "tipo": "Indirecto", "estrella": 3,
     "sms": "Sí", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Cumple con una parte de las SMS al conocer y gestionar información de trabajadores.",
     "demo": "Canal de denuncias y código de ética para DDHH.",
     "falta": "Información debe usarse para garantizar derechos según Carta Internacional de DDHH y OIT."},
    {"id": 5, "dim": "Trabajadores", "criterio": "Diseño de plan de desarrollo y bienestar", "tipo": "Directo", "estrella": 3,
     "sms": "Sí", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Se vincula directamente con las SMS al establecer iniciativas de bienestar.",
     "demo": "Plan de desarrollo y bienestar que incluya prevención de delitos.", "falta": ""},
    {"id": 6, "dim": "Trabajadores", "criterio": "Formación continua", "tipo": "Indirecto", "estrella": 5,
     "sms": "No", "cs1": "Sí", "cs2": "Sí", "n1": "Sí", "n2": "Sí", "n3": "Sí", "n4": "Sí", "n5": "Sí", "n6": "Sí",
     "justif": "Se vincula indirectamente con los 6 OM si las capacitaciones incluyen temáticas T-MAS.",
     "demo": "Certificados de capacitación en temas ambientales.",
     "falta": "Capacitaciones deben incluir específicamente los CTS de la T-MAS."},
    {"id": 7, "dim": "Trabajadores", "criterio": "Diversidad e inclusión laboral", "tipo": "Directo", "estrella": 1,
     "sms": "Sí", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Garantiza respeto a los DDHH, alineándose con las SMS.",
     "demo": "Medidas para evitar discriminación y canales de denuncias.", "falta": ""},
    {"id": 8, "dim": "Trabajadores", "criterio": "Participación de mujeres en la industria", "tipo": "Directo", "estrella": 1,
     "sms": "Sí", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Garantiza respeto a los DDHH de cada mujer.",
     "demo": "Participación femenina y medidas de fomento.", "falta": ""},
    {"id": 9, "dim": "Trabajadores", "criterio": "Inclusión de personas con discapacidad", "tipo": "Directo", "estrella": 1,
     "sms": "Sí", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Garantiza respeto a los DDHH.",
     "demo": "Medidas para evitar discriminación.", "falta": ""},
    {"id": 10, "dim": "SST", "criterio": "Gestión de objetivos e indicadores de seguridad", "tipo": "Indirecto", "estrella": 1,
     "sms": "Sí", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Garantiza DDHH en salud y seguridad laboral.",
     "demo": "Medidas de salud laboral según OIT.",
     "falta": "Indicadores deben alinearse con estándares OIT de la T-MAS."},
    {"id": 11, "dim": "SST", "criterio": "Gestión de desarrollo y capacitación de trabajadores", "tipo": "Indirecto", "estrella": 1,
     "sms": "Sí", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Garantiza DDHH mediante capacitación en seguridad.",
     "demo": "Medidas de salud laboral según OIT.",
     "falta": "Capacitaciones deben incluir riesgos ambientales y climáticos de la T-MAS."},
    {"id": 12, "dim": "Cadena de Valor", "criterio": "Desarrollo de proveedores, contratistas y especialistas", "tipo": "Indirecto", "estrella": 1,
     "sms": "Sí", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Proveedores pueden usar canales de denuncia de la empresa.",
     "demo": "Canales de denuncia abiertos a toda la cadena de valor.",
     "falta": "Evaluar proveedores en criterios de sostenibilidad T-MAS y exigir reportes de GEI."},
    {"id": 13, "dim": "Cadena de Valor", "criterio": "Gestión de la sostenibilidad con la cadena de valor", "tipo": "Directo", "estrella": 5,
     "sms": "No", "cs1": "Sí", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Se alinea directamente con CS del OM1 al orientar criterios de contratación a GEI.",
     "demo": "Certificados, reportes de emisiones de proveedores.", "falta": ""},
    {"id": 14, "dim": "Comunidad", "criterio": "Estrategia de relacionamiento comunitario", "tipo": "Indirecto", "estrella": 1,
     "sms": "Sí", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "La empresa puede contar con canales de denuncia de la comunidad.",
     "demo": "Canales de denuncia para abordar necesidades de la comunidad.",
     "falta": "Canales deben incluir temas ambientales y gestionar reclamos de impacto ambiental."},
    {"id": 15, "dim": "Medioambiente", "criterio": "Gestión de acciones sostenibles del espacio de trabajo", "tipo": "Directo", "estrella": 1,
     "sms": "Sí", "cs1": "Sí", "cs2": "No", "n1": "No", "n2": "No", "n3": "Sí", "n4": "Sí", "n5": "No", "n6": "No",
     "justif": "Medidas para gestionar residuos, energía y agua alineadas con SMS, CS y NHDS.",
     "demo": "Respaldo de medidas de gestión de residuos, consumo hídrico y ahorro energético.", "falta": ""},
    {"id": 16, "dim": "Medioambiente", "criterio": "Gestión de residuos", "tipo": "Directo", "estrella": 1,
     "sms": "No", "cs1": "No", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "Sí", "n5": "No", "n6": "No",
     "justif": "Permite gestionar residuos alineado a NHDS.",
     "demo": "Respaldo de medidas de gestión de residuos.", "falta": ""},
    {"id": 17, "dim": "Medioambiente", "criterio": "Consumo de energía en las instalaciones", "tipo": "Directo", "estrella": 1,
     "sms": "No", "cs1": "Sí", "cs2": "No", "n1": "Sí", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Permite reducir consumo eléctrico y emisiones de GEI.",
     "demo": "Respaldo del consumo energético que demuestre disminución.", "falta": ""},
    {"id": 18, "dim": "Medioambiente", "criterio": "Emisiones GEI corporativas", "tipo": "Directo", "estrella": 1,
     "sms": "No", "cs1": "Sí", "cs2": "No", "n1": "Sí", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "La medición permite gestionar emisiones de GEI. Directamente relacionado con OM1.",
     "demo": "Medición de huella de carbono y verificación por terceros.", "falta": ""},
    {"id": 19, "dim": "Medioambiente", "criterio": "Emisiones: Huella de carbono de productos", "tipo": "Directo", "estrella": 1,
     "sms": "No", "cs1": "Sí", "cs2": "No", "n1": "Sí", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Medición de huella de carbono de productos, directamente relacionado con OM1.",
     "demo": "Medición de huella de carbono de productos.", "falta": ""},
    {"id": 20, "dim": "Medioambiente", "criterio": "Oferta de productos con atributos de sustentabilidad", "tipo": "Directo", "estrella": 2,
     "sms": "No", "cs1": "Sí", "cs2": "No", "n1": "Sí", "n2": "No", "n3": "No", "n4": "Sí", "n5": "No", "n6": "No",
     "justif": "Desde la Estrella 2 se alinea con CTS de CS del OM1 para fabricación de productos.",
     "demo": "Fichas técnicas, DAP o certificaciones de producto.", "falta": ""},
    {"id": 21, "dim": "Medioambiente", "criterio": "Adaptación al cambio climático", "tipo": "Directo", "estrella": 2,
     "sms": "No", "cs1": "No", "cs2": "Sí", "n1": "No", "n2": "Sí", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "Desde la estrella 2 se cumple con NHDS de Adaptación y desde la 4 con CS a la adaptación.",
     "demo": "Evaluación y priorización de riesgos climáticos.", "falta": ""},
    {"id": 22, "dim": "Innovación y Productividad", "criterio": "Innovación", "tipo": "Directo", "estrella": 1,
     "sms": "No", "cs1": "Sí", "cs2": "Sí", "n1": "No", "n2": "No", "n3": "No", "n4": "Sí", "n5": "No", "n6": "No",
     "justif": "La innovación puede contribuir a reducción de emisiones y economía circular.",
     "demo": "Documentación de proyectos de innovación con impacto ambiental.", "falta": ""},
    {"id": 23, "dim": "Innovación y Productividad", "criterio": "Transformación digital", "tipo": "Indirecto", "estrella": 1,
     "sms": "No", "cs1": "Sí", "cs2": "No", "n1": "No", "n2": "No", "n3": "No", "n4": "No", "n5": "No", "n6": "No",
     "justif": "La transformación digital puede contribuir indirectamente a eficiencia energética.",
     "demo": "Proyectos digitales que demuestren reducción de consumo.",
     "falta": "Demostrar reducción medible de emisiones de GEI o consumo energético."},
]

def crear_template_matriz():
    """Crea el template de la Matriz de Vinculación"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriz Vinculación"
    
    # ---- Título ----
    ws.merge_cells('A1:Q1')
    ws['A1'] = '📋 MATRIZ DE VINCULACIÓN SELLOPRO ↔ T-MAS'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:Q2')
    ws['A2'] = 'Por favor VALIDEN y CORRIJAN la información. Las celdas amarillas requieren especial atención.'
    ws['A2'].font = Font(size=10, italic=True, color="F59E0B")
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 25
    
    # ---- Headers ----
    headers = [
        ('A', '#', 5),
        ('B', 'Dimensión', 20),
        ('C', 'Criterio SelloPRO', 35),
        ('D', 'Tipo Vínculo', 15),
        ('E', 'Estrella Mínima', 15),
        ('F', 'SMS', 8),
        ('G', 'CS OM1\nMitigación', 12),
        ('H', 'CS OM2\nAdaptación', 12),
        ('I', 'NHDS OM1', 10),
        ('J', 'NHDS OM2', 10),
        ('K', 'NHDS OM3\nAgua', 10),
        ('L', 'NHDS OM4\nResiduos', 10),
        ('M', 'NHDS OM5\nContam.', 10),
        ('N', 'NHDS OM6\nBiodiv.', 10),
        ('O', 'Justificación\n¿POR QUÉ se relaciona?', 45),
        ('P', '¿Cómo lo demuestra\nla empresa?', 40),
        ('Q', '(Solo Indirectos)\n¿Qué falta para Directo?', 45),
    ]
    
    row = 4
    for col_letter, header, width in headers:
        cell = ws[f'{col_letter}{row}']
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = DARK_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[row].height = 40
    
    # ---- Datos pre-llenados ----
    for i, c in enumerate(CRITERIOS_ACTUALES):
        r = row + 1 + i
        data = [
            c['id'], c['dim'], c['criterio'], c['tipo'], c['estrella'],
            c['sms'], c['cs1'], c['cs2'], c['n1'], c['n2'], c['n3'], c['n4'], c['n5'], c['n6'],
            c['justif'], c['demo'], c['falta']
        ]
        
        for j, val in enumerate(data):
            col = get_column_letter(j + 1)
            cell = ws[f'{col}{r}']
            cell.value = val
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Colorear según tipo
            if j == 3:  # Tipo vínculo
                if val == 'Directo':
                    cell.fill = LIGHT_GREEN_FILL
                elif val == 'Indirecto':
                    cell.fill = YELLOW_FILL
            
            # Centrar las columnas de Sí/No
            if 5 <= j <= 13:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[r].height = 50
    
    # Agregar filas vacías para criterios nuevos
    empty_start = row + 1 + len(CRITERIOS_ACTUALES)
    for i in range(10):
        r = empty_start + i
        for j in range(17):
            col = get_column_letter(j + 1)
            cell = ws[f'{col}{r}']
            cell.border = THIN_BORDER
            cell.fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    
    # ---- Validaciones de datos ----
    # Tipo vínculo
    dv_tipo = DataValidation(type="list", formula1='"Directo,Indirecto,Sin Vínculo"', allow_blank=True)
    dv_tipo.error = "Debe ser: Directo, Indirecto o Sin Vínculo"
    dv_tipo.errorTitle = "Valor no válido"
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f'D5:D{empty_start + 9}')
    
    # Estrellas
    dv_estrellas = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
    dv_estrellas.error = "Debe ser un número entre 1 y 5"
    ws.add_data_validation(dv_estrellas)
    dv_estrellas.add(f'E5:E{empty_start + 9}')
    
    # Sí/No
    dv_sino = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
    ws.add_data_validation(dv_sino)
    for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        dv_sino.add(f'{col}5:{col}{empty_start + 9}')
    
    # ---- Hoja de instrucciones ----
    ws_instr = wb.create_sheet("Instrucciones", 0)
    instrucciones = [
        ("📋 INSTRUCCIONES PARA LLENAR LA MATRIZ", TITLE_FONT, None),
        ("", None, None),
        ("CONTEXTO:", BOLD_FONT, None),
        ("Esta matriz establece la relación entre cada criterio del SelloPRO y los objetivos", NORMAL_FONT, None),
        ("de la Taxonomía de Actividades Medioambientalmente Sostenibles de Chile (T-MAS).", NORMAL_FONT, None),
        ("Los datos pre-llenados fueron extraídos del Excel 'base suministro.xlsx' que nos entregaron.", NORMAL_FONT, None),
        ("", None, None),
        ("QUÉ NECESITAMOS DE USTEDES:", BOLD_FONT, None),
        ("1. VALIDAR que cada criterio tenga el tipo de vínculo correcto (Directo/Indirecto/Sin Vínculo)", NORMAL_FONT, None),
        ("2. VALIDAR que la estrella mínima sea correcta para cada criterio", NORMAL_FONT, None),
        ("3. VALIDAR los objetivos T-MAS que impacta cada criterio (columnas F a N)", NORMAL_FONT, None),
        ("4. MEJORAR la columna 'Justificación' con lenguaje simple y motivador", NORMAL_FONT, None),
        ("5. MEJORAR la columna 'Cómo lo demuestra' con documentos/evidencias específicas", NORMAL_FONT, None),
        ("6. COMPLETAR la columna 'Qué falta para Directo' para todos los INDIRECTOS", NORMAL_FONT, None),
        ("7. AGREGAR criterios que falten en las filas amarillas vacías", NORMAL_FONT, None),
        ("", None, None),
        ("DEFINICIONES:", BOLD_FONT, None),
        ("• Directo: El criterio SelloPRO cumple directamente con el requisito T-MAS.", NORMAL_FONT, None),
        ("• Indirecto: El criterio SelloPRO contribuye parcialmente. Hay acciones adicionales para cumplir.", NORMAL_FONT, None),
        ("• Sin Vínculo: El criterio SelloPRO no tiene relación con la T-MAS.", NORMAL_FONT, None),
        ("• Estrella Mínima: Desde qué nota (1-5) el criterio empieza a vincularse con la T-MAS.", NORMAL_FONT, None),
        ("", None, None),
        ("OBJETIVOS T-MAS:", BOLD_FONT, None),
        ("• SMS = Salvaguardas Mínimas Sociales (DDHH, ética, canal de denuncias)", NORMAL_FONT, None),
        ("• CS OM1 = Contribución Sustancial a Mitigación del Cambio Climático", NORMAL_FONT, None),
        ("• CS OM2 = Contribución Sustancial a Adaptación al Cambio Climático", NORMAL_FONT, None),
        ("• NHDS OM1 = No Hacer Daño Significativo a Mitigación", NORMAL_FONT, None),
        ("• NHDS OM2 = No Hacer Daño Significativo a Adaptación", NORMAL_FONT, None),
        ("• NHDS OM3 = No Hacer Daño Significativo al Uso de Agua", NORMAL_FONT, None),
        ("• NHDS OM4 = No Hacer Daño Significativo a Economía Circular (Residuos)", NORMAL_FONT, None),
        ("• NHDS OM5 = No Hacer Daño Significativo a Prevención de Contaminación", NORMAL_FONT, None),
        ("• NHDS OM6 = No Hacer Daño Significativo a Protección de Biodiversidad", NORMAL_FONT, None),
        ("", None, None),
        ("⚠️ IMPORTANTE:", BOLD_FONT, None),
        ("La columna 'Qué falta para Directo' es CLAVE para los socios. Es lo que les muestra", NORMAL_FONT, None),
        ("qué acciones concretas tomar para mejorar su alineación. Sea ESPECÍFICO y MOTIVADOR.", NORMAL_FONT, None),
        ("", None, None),
        ("📅 PLAZO: Por favor devolver este archivo llenado dentro de 1 semana.", BOLD_FONT, None),
    ]
    
    ws_instr.column_dimensions['A'].width = 90
    for i, (texto, font, fill) in enumerate(instrucciones):
        cell = ws_instr[f'A{i+1}']
        cell.value = texto
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
    
    wb.save('TEMPLATE_01_matriz_vinculacion.xlsx')
    print("✅ Template 1 creado: TEMPLATE_01_matriz_vinculacion.xlsx")


def crear_template_estrellas():
    """Crea el template de Descripción de Estrellas"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Descripción Estrellas"
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = '⭐ DESCRIPCIÓN DE ESTRELLAS POR CRITERIO SELLOPRO'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:G2')
    ws['A2'] = '¿Qué significa tener 1, 2, 3, 4 o 5 estrellas en cada criterio?'
    ws['A2'].font = Font(size=10, italic=True, color="475569")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = [
        ('A', '#', 5),
        ('B', 'Criterio', 35),
        ('C', '⭐ (1 estrella)', 30),
        ('D', '⭐⭐ (2 estrellas)', 30),
        ('E', '⭐⭐⭐ (3 estrellas)', 30),
        ('F', '⭐⭐⭐⭐ (4 estrellas)', 30),
        ('G', '⭐⭐⭐⭐⭐ (5 estrellas)', 30),
    ]
    
    row = 4
    for col_letter, header, width in headers:
        cell = ws[f'{col_letter}{row}']
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = DARK_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = width
    
    # Datos pre-llenados (solo los que tenemos)
    estrellas_data = {
        1: {"1": "No cuenta con iniciativas ni prácticas en sostenibilidad",
            "2": "Existe un Plan con prácticas o iniciativas recurrentes",
            "3": "La sostenibilidad es liderada por un encargado, área o comité",
            "4": "Cuenta con una estrategia de sostenibilidad integrada en el negocio",
            "5": "La estrategia se construyó en base a un estudio de materialidad"},
        2: {"1": "Comité de Gestión interna que se reúne periódicamente",
            "2": "Asesor externo que guía en temas normativos y financieros",
            "3": "Directorio que vela por los objetivos estratégicos",
            "4": "Directorio compuesto por un grupo diverso",
            "5": "El directorio se capacita en temas de sostenibilidad"},
        3: {"1": "Reglamento interno con misión, visión y valores",
            "2": "Difunde su misión, visión y valores con la organización",
            "3": "Código de Ética y capacita a los miembros",
            "4": "Canal de denuncias que permite confidencialidad",
            "5": "Canal gestionado por un comité de ética con roles asignados"},
    }
    
    for i, c in enumerate(CRITERIOS_ACTUALES):
        r = row + 1 + i
        ws[f'A{r}'] = c['id']
        ws[f'A{r}'].font = NORMAL_FONT
        ws[f'A{r}'].border = THIN_BORDER
        ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'B{r}'] = c['criterio']
        ws[f'B{r}'].font = BOLD_FONT
        ws[f'B{r}'].border = THIN_BORDER
        ws[f'B{r}'].alignment = Alignment(vertical='center', wrap_text=True)
        
        # Si tenemos datos de estrellas, llenar
        if c['id'] in estrellas_data:
            for star in range(1, 6):
                col = get_column_letter(star + 2)
                cell = ws[f'{col}{r}']
                cell.value = estrellas_data[c['id']].get(str(star), '')
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        else:
            for star in range(1, 6):
                col = get_column_letter(star + 2)
                cell = ws[f'{col}{r}']
                cell.value = ''
                cell.fill = YELLOW_FILL
                cell.border = THIN_BORDER
        
        ws.row_dimensions[r].height = 45
    
    wb.save('TEMPLATE_02_descripcion_estrellas.xlsx')
    print("✅ Template 2 creado: TEMPLATE_02_descripcion_estrellas.xlsx")


def crear_template_objetivos():
    """Crea el template de Objetivos T-MAS"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Objetivos T-MAS"
    
    # Título
    ws.merge_cells('A1:K1')
    ws['A1'] = '🎯 INFORMACIÓN OFICIAL DE OBJETIVOS T-MAS'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:K2')
    ws['A2'] = 'Descripciones, requisitos y documentos de cada objetivo de la Taxonomía. VALIDAR con fuentes oficiales.'
    ws['A2'].font = Font(size=10, italic=True, color="475569")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = [
        ('A', 'Código', 12),
        ('B', 'Nombre', 25),
        ('C', 'Descripción\n(lenguaje simple)', 45),
        ('D', 'Requisito 1', 35),
        ('E', 'Requisito 2', 35),
        ('F', 'Requisito 3', 35),
        ('G', 'Requisito 4', 35),
        ('H', 'Documento 1', 25),
        ('I', 'Documento 2', 25),
        ('J', 'Documento 3', 25),
        ('K', 'Referencia Oficial', 30),
    ]
    
    row = 4
    for col_letter, header, width in headers:
        cell = ws[f'{col_letter}{row}']
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = DARK_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[row].height = 35
    
    # Datos pre-llenados
    objetivos = [
        {"id": "SMS", "nombre": "Compromiso Social y Ético",
         "desc": "Tu empresa demuestra que opera de forma ética y responsable.",
         "req": ["Procedimientos para cumplir leyes laborales, ambientales y tributarias",
                 "Canal de denuncias seguro para trabajadores",
                 "Respeto a derechos de trabajadores",
                 "Sin multas o sanciones graves en últimos 5 años"],
         "docs": ["Código de ética", "Reglamento interno", "Canal de denuncias"]},
        {"id": "CS_OM1", "nombre": "Reducción de Emisiones",
         "desc": "Tu empresa toma acciones concretas para reducir su huella de carbono.",
         "req": ["Mides cuánta energía consume tu empresa",
                 "Implementas medidas para reducir el consumo",
                 "Conoces tu huella de carbono",
                 "Tienes un plan para seguir reduciendo emisiones"],
         "docs": ["Medición de huella de carbono", "Plan de eficiencia energética", "Facturas de energía"]},
        {"id": "CS_OM2", "nombre": "Preparación ante el Clima",
         "desc": "Tu empresa está preparada para los efectos del cambio climático.",
         "req": ["Identificas qué riesgos climáticos pueden afectar tu negocio",
                 "Evalúas qué tan vulnerable es tu operación",
                 "Tienes medidas para adaptarte",
                 "Tu plan no perjudica a vecinos ni al medio ambiente"],
         "docs": ["Análisis de riesgos climáticos", "Plan de contingencia", "Medidas de adaptación"]},
        {"id": "NHDS_OM1", "nombre": "Sin Daño al Clima",
         "desc": "Tu operación no genera un impacto negativo significativo en el clima.",
         "req": ["Cumples con permisos ambientales",
                 "Conoces y controlas las emisiones",
                 "No generas contaminación excesiva", ""],
         "docs": ["Permisos ambientales al día", "Registro de emisiones", ""]},
        {"id": "NHDS_OM2", "nombre": "Sin Daño a la Adaptación",
         "desc": "Tu actividad no impide que otros se adapten al cambio climático.",
         "req": ["Tu operación no aumenta la vulnerabilidad de la zona",
                 "Cumples con la normativa ambiental",
                 "No afectas negativamente a vecinos o ecosistemas", ""],
         "docs": ["Permisos de construcción/operación", "Cumplimiento normativo", ""]},
        {"id": "NHDS_OM3", "nombre": "Cuidado del Agua",
         "desc": "Usas el agua de forma responsable.",
         "req": ["Mides cuánta agua consume tu empresa",
                 "Tienes medidas para reducir el consumo",
                 "No contaminas fuentes de agua", ""],
         "docs": ["Registro de consumo de agua", "Plan de uso eficiente del agua", ""]},
        {"id": "NHDS_OM4", "nombre": "Gestión de Residuos",
         "desc": "Manejas bien tus residuos.",
         "req": ["Separas los residuos",
                 "Reciclas lo que se puede",
                 "Residuos peligrosos van a lugares autorizados",
                 "Cumples con la Ley REP si aplica"],
         "docs": ["Plan de gestión de residuos", "Certificados de disposición", "Registros de reciclaje"]},
        {"id": "NHDS_OM5", "nombre": "Control de Contaminación",
         "desc": "Tu empresa controla sus emisiones al aire, ruidos y otros contaminantes.",
         "req": ["Cumples con los límites de emisión",
                 "Monitoreas tus emisiones regularmente",
                 "Tienes medidas para reducir la contaminación", ""],
         "docs": ["Mediciones de emisiones", "Permisos ambientales", "Plan de control de contaminación"]},
        {"id": "NHDS_OM6", "nombre": "Protección de la Naturaleza",
         "desc": "Tu operación respeta los ecosistemas y la biodiversidad.",
         "req": ["No operas en áreas protegidas sin autorización",
                 "Identificas especies protegidas en tu zona",
                 "Minimizas tu impacto en la naturaleza", ""],
         "docs": ["Evaluación de impacto ambiental", "Permisos de operación", "Plan de manejo ambiental"]},
    ]
    
    for i, obj in enumerate(objetivos):
        r = row + 1 + i
        ws[f'A{r}'] = obj['id']
        ws[f'A{r}'].font = BOLD_FONT
        ws[f'A{r}'].border = THIN_BORDER
        ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'B{r}'] = obj['nombre']
        ws[f'B{r}'].font = BOLD_FONT
        ws[f'B{r}'].border = THIN_BORDER
        ws[f'B{r}'].alignment = Alignment(vertical='center', wrap_text=True)
        
        ws[f'C{r}'] = obj['desc']
        ws[f'C{r}'].font = NORMAL_FONT
        ws[f'C{r}'].border = THIN_BORDER
        ws[f'C{r}'].alignment = Alignment(vertical='center', wrap_text=True)
        
        for j, req in enumerate(obj['req']):
            col = get_column_letter(4 + j)
            cell = ws[f'{col}{r}']
            cell.value = req
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        for j, doc in enumerate(obj['docs']):
            col = get_column_letter(8 + j)
            cell = ws[f'{col}{r}']
            cell.value = doc
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        ws[f'K{r}'] = ''
        ws[f'K{r}'].fill = YELLOW_FILL
        ws[f'K{r}'].border = THIN_BORDER
        
        ws.row_dimensions[r].height = 60
    
    wb.save('TEMPLATE_03_objetivos_tmas.xlsx')
    print("✅ Template 3 creado: TEMPLATE_03_objetivos_tmas.xlsx")


if __name__ == '__main__':
    print("=" * 60)
    print("  Generando templates para consultores...")
    print("=" * 60)
    crear_template_matriz()
    crear_template_estrellas()
    crear_template_objetivos()
    print()
    print("=" * 60)
    print("  ✅ ¡3 templates generados exitosamente!")
    print("=" * 60)
    print()
    print("Archivos creados:")
    print("  📄 TEMPLATE_01_matriz_vinculacion.xlsx")
    print("  📄 TEMPLATE_02_descripcion_estrellas.xlsx")
    print("  📄 TEMPLATE_03_objetivos_tmas.xlsx")
    print()
    print("Enviar estos 3 archivos a los consultores para que")
    print("VALIDEN y COMPLETEN la información.")

